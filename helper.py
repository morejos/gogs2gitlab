import time
import pyperclip
import pandas as pd
from datetime import datetime
from pytz import timezone
from constants import *
from openpyxl import load_workbook
from seleniumbase import Driver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains
from selenium.common.exceptions import NoSuchElementException

class HelperClass:
    def createProject(self, driver, counter, gogsLink, projName):
        # "New Project"
        driver.find_element(By.LINK_TEXT, "New project").click()
        time.sleep(2)

        # "Import Project"
        driver.find_element(By.XPATH, "//div[@class='new-namespace-panel-wrapper gl-display-inline-block gl-float-left gl-px-3 gl-mb-5'][3]").click()
        time.sleep(2)

        # "Repository by URL"
        driver.find_element(By.XPATH, "//button[@data-platform='repo_url']").click()
        time.sleep(2)

        # Enter Git repo URL
        driver.find_element(By.ID, "project_import_url").send_keys(gogsLink)
        time.sleep(2)

        # Clear out project name & enter from Excel list (for safe measures)
        projectNameBox = driver.find_element(By.ID, "project_name")
        projectNameBox.clear()
        projectNameBox.send_keys(projName)
        time.sleep(2)

        # Make it public
        driver.find_element(By.XPATH, "//label[@for='project_visibility_level_20']").click()
        time.sleep(2)

        # "Create Project" 
        driver.find_element(By.XPATH, "//button[@data-qa-selector='project_create_button']").click()
        cloneBtn = WebDriverWait(driver, 20).until(EC.element_to_be_clickable((By.ID, "clone-dropdown")))

        # Verify successfull creation of project, otherwise if we missed this project already being created, skip to next project
        if ((SUCCESS_MESSAGE in driver.page_source) == True):
            print(">>>>> (1/2) Successfully imported project", projName)
        else:
            print("[ERROR] Failed to create project", projName, "moving onto next project.\n")
            driver.find_element(By.LINK_TEXT, "QBITAutomation").click()
            time.sleep(1.5)
            counter += 1
            return
        
        # "Clone"
        cloneBtn.click()
        time.sleep(2)

        # "Copy URL"
        driver.find_element(By.XPATH, "//button[@data-clipboard-target='#http_project_clone']").click()
        time.sleep(2)

        # Write the gitlab link to the excel sheet
        HelperClass.writeToExcel(self, counter, projName, False)

        # Navigate back to QBITAutomation group home page
        driver.find_element(By.LINK_TEXT, "QBITAutomation").click()
        time.sleep(2)

    def writeToExcel(self, counter, projName, doesItExists):
        wb = load_workbook(filename = MYXSX)
        sheet = wb['Sheet1']
        
        if (doesItExists == False):
            # Write the gitlab link to the excel sheet
            if (sheet.cell(row=counter, column=PROJECT_TITLE).value == str(projName)[2:-2]): # Verify here the correct row is selected
                sheet.cell(row=counter, column=GITLAB_LINK).value = pyperclip.paste()
                wb.save(MYXSX)
                print(">>>>> (2/2) Successfully updated excel sheet for project", projName, "\n")
            else:
                print("[ERROR] Unable to add GitLab link", projName, "to excel sheet\n")
        else:
            sheet.cell(row=counter, column=GITLAB_LINK).value = "THIS PROJECT ALREADY EXISTS ON GITLAB."
            wb.save(MYXSX)

    def writeInfoToExcel(self, counter, projName, reason):
        wb = load_workbook(filename = MYXSX)
        sheet = wb['Sheet1']
        
        if (reason == "DNE"):
            print("[ERROR]", projName, "does not exist, updating excel sheet.")
            sheet.cell(row=counter, column=INFO_COLUMN).value = "Project not found on Jenkins."
            wb.save(MYXSX)
            print(">>>>> Successfully updated excel sheet for missing project:", projName, "\n")
        elif (reason == "FAIL"):
            sheet.cell(row=counter, column=SECOND_INFO_COLUMN).value = "Failed to configure GitLab and/or Jenkins, please review."
            wb.save(MYXSX)
        elif (reason == "DONE"):
            sheet.cell(row=counter, column=SECOND_INFO_COLUMN).value = "Successfully configured GitLab + Jenkins."
            wb.save(MYXSX)
        else:
            print("[ERROR] Unable to update project info message for:", projName, "to excel sheet\n")

    def loginGitLab(self, driver):
        # "Sign in"
        driver.find_element(By.LINK_TEXT, "Sign in").click()
        time.sleep(0.5)

        # Enter credentials
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.NAME, "username"))).send_keys(USER)
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.NAME, "password"))).send_keys(PSWD)
        time.sleep(0.5)

        # "Sign in"
        driver.find_element(By.NAME, "commit").click()
        time.sleep(1)

    def loginJenkins(self, driver):
        print(">>>>> Completed importing projects to GitLab, now updating Jenkins configurations...\n")
        
        tf = True
        while tf:
            try:
                # Enter credentials
                WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.NAME, "j_username"))).send_keys(USER)
                WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.NAME, "j_password"))).send_keys(PSWD)
                time.sleep(0.5)

                # "Sign in"
                signInBtn = driver.find_element(By.NAME, "Submit")
                signInBtn.click()
                time.sleep(5)
                tf = False
            except Exception:
                driver.refresh()
                continue

    def initializeChromeDriver(self, URL):
        driver = Driver(uc=True)
        driver.maximize_window()
        driver.get(URL)
        time.sleep(1)
        return driver

    def returnDateAndTime(self):
        # Grab today's date and time (EST) + format to match Jenkins
        dateTime = datetime.now(timezone('US/Eastern'))
        return dateTime.strftime("%b %d, %Y, %#I:%M %p ")

    def modifyJenkinsProject(self, projectNamesList):
        # Helps us keep track of what excel row we should be on
        counter = 2

        # Open a second Chrome window
        driver = HelperClass.initializeChromeDriver(self, JENKINS_URL)

        # Login to Jenkins
        HelperClass.loginJenkins(self, driver)

        # Grab list of projects from Jenkins home page
        jenkins_projects = driver.find_elements(By.XPATH, "//a[@class='jenkins-table__link model-link inside']")
        
        # Read the new GitLab links for future use
        gitLabLinks = pd.read_excel(MYXSX, header=None, skiprows=1, usecols="C").values.tolist()

        # Loop through all the projects in the Excel sheet so we may update Jenkins configs
        for gitLink, projName in zip(gitLabLinks, projectNamesList):
            
            # Re-assign projects list on Jenkins & re-set the check variable
            jenkins_projects = driver.find_elements(By.XPATH, "//a[@class='jenkins-table__link model-link inside']")
            projectExists = False

            # Clean up string variable values
            projName = str(projName)[2:-2]
            gitLink = str(gitLink)[2:-2]

            # Verify the project exists on Jenkins, if so click on it
            for jProject in jenkins_projects:
                if (projName == jProject.text):
                    projectExists = True
                    break
            
            # If project exists, click on it so we can update the configuration
            if (projectExists == True):
                print(">>>>> (1/3) Successfully found Jenkins project:", projName)
                element = driver.find_element(By.XPATH, "//a[@href='job/"+projName+"/']")
                driver.execute_script("arguments[0].click();", element)
                time.sleep(1.5)

                # First grab the build time so we know how long to wait to avoid errors
                actionChains = ActionChains(driver)
                actionChains.double_click(driver.find_element(By.LINK_TEXT, 'master')).perform()
                time.sleep(1)

                # Verify we can see "Pipeline master" to make sure we are on right page & grab average build time for later use
                WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, "//div[@id='pipeline-box']")))
                avgBuildTime = (HelperClass.getBuildTime(self, driver) + 20)
                time.sleep(1.5)

                # Nav back one page
                driver.back()
                time.sleep(2)

                # "Configure"
                WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.LINK_TEXT, "Configure"))).click()
                time.sleep(2) # time.sleep(avgBuildTime)

                # Remove old first link w/ new one
                element = driver.find_element(By.NAME, "_.remote")
                element.clear()
                element.send_keys(gitLink)
                time.sleep(1)

                # Update below properties (if project has them)
                try:
                    # Update configure repo broser from Gogs > GitLab
                    selectDropDown = Select(driver.find_element(By.XPATH, "//select[@class='jenkins-select__input dropdownList']"))
                    selectDropDown.select_by_value('10')
                    time.sleep(1)
                    
                    # Update configure repo browser URL
                    element2 = driver.find_element(By.NAME, "_.repoUrl")
                    element2.clear()
                    element2.send_keys(gitLink)
                    time.sleep(1)
                except:
                    print(">>>>> Proceeding w/ remaining steps as no secondary Git entry text box for:", projName)
                    time.sleep(1)

                # "Save" & wait depending on which page of Jenkins we go to
                driver.find_element(By.XPATH, "//button[@type='submit']").click()
                if ((DISABLE_PIPELINE in driver.page_source) == True): # "Disable Multibranch Pipeline" page
                    time.sleep(2)
                elif ((VIEW_PLAIN_TEXT in driver.page_source) == True): # "Scan Multibranch Pipeline" page
                    driver.find_element(By.XPATH, "//div[@id='tasks']/div[1]").click()
                    time.sleep(2)
                else:
                    time.sleep(avgBuildTime)

                # "master" branch (We have to double click this element thus action chains)
                actionChains = ActionChains(driver)
                actionChains.double_click(driver.find_element(By.LINK_TEXT, 'master')).perform()
                time.sleep(2.5)

                # Build updated project w/ updated Jenkins configs & verify
                HelperClass.buildAndVerify(self, driver, avgBuildTime, projName, counter)
                counter += 1
                time.sleep(1)
                
                # Setup for next project
                HelperClass.navBackToDashboard(self, driver)
            else:
                HelperClass.writeInfoToExcel(self, counter, projName, "DNE")
                print("[ERROR] Jenkins project:", projName, "could not be found, excel sheet has been updated.\n")
                counter += 1

                # Setup for next project
                HelperClass.navBackToDashboard(self, driver)
        # End
        driver.close()

    def navBackToDashboard(self, driver):
        driver.find_element(By.LINK_TEXT, "Dashboard").click()

         # Make it wait until we see this
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.LINK_TEXT, "Fav")))

    def getBuildTime(self, driver):
        # Grab the average build time, this will be used later
        buildTime = driver.find_element(By.XPATH, "//td[@class='stage-total-2']").text

        # Now check if we are dealing w/ a single second, multi-second, or over minute build time
        buildTimeLen = len(buildTime)

        if (buildTimeLen < 3):
            # Single second build time
            return int(float(buildTime[0:1]))
        elif (buildTimeLen == 3):
            # Multi-second build time
            return int(float(buildTime[0:2]))
        elif (7 <= buildTimeLen <= 8):
            # Over a minute build time
            runTimeMin = int(float(buildTime[0:1]))
            runTimeSec = buildTime[:len(buildTime) - 1]
            runTimeSec = int(float(runTimeSec[len(runTimeSec) - 2:]))
        
            if (runTimeMin == 1):
                return 60 + runTimeSec
            elif (runTimeMin == 2):
                return 120 + runTimeSec
            else:
                return 180 + runTimeSec
        else:
            # In case project has no build history
            return 180

    def buildAndVerify(self, driver, avgBuildTime, projName, counter):
        # Grab the number of current builds (otherwise make an empty list if we have no build history)
        try:
            builds = driver.find_elements(By.XPATH, "//td[@class='build-row-cell']")
        except NoSuchElementException:
            builds = []

        # Build the Jenkins project now using GitLab & wait the average build time
        driver.find_element(By.XPATH, "//div[@id='tasks']/div[3]").click()
        time.sleep(avgBuildTime)

        # Check the number of builds again verifying a new build was added
        builds2 = driver.find_elements(By.XPATH, "//td[@class='build-row-cell']")

        # Compare both lists verifying new lists is +1 on old list
        if (len(builds) < len(builds2)):
            # Refresh the screen so we may see new statistics
            driver.refresh()

            # Re-initialize build elements since page was refreshed: myTime = HelperClass.returnDateAndTime(self)
            builds2 = driver.find_elements(By.XPATH, "//td[@class='build-row-cell']")
            latestBuildTime = builds2[0].text

            # Parse each time accordingly
            latestBuildID = latestBuildTime.partition('\n')[0]
            lastSuccessfulBuild = driver.find_element(By.XPATH, "//a[@href='lastSuccessfulBuild/']").text

            # Verify we see latest build ID as "Last Sucessful Build"
            if (latestBuildID in lastSuccessfulBuild):
                print(">>>>> (2/3) Successfully validated build for:", projName)
                print(">>>>> (3/3) Completed Jenkins configuration for:", projName, "\n")
                HelperClass.writeInfoToExcel(self, counter, projName, "DONE")
            else:
                print("[ERROR] Failed to validate build status for project:", projName, ", please check afterwards")
                HelperClass.writeInfoToExcel(self, counter, projName, "FAIL")
        else:
            print("[ERROR] Failed to check for updated Jenkins build for", projName, ", please check afterwards.")